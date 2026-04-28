"""
verification_export.py — Export verification plans to Excel and Markdown summary.

Usage:
    python verification_export.py excel \\
        --plan VERIFICATION.md \\
        --results TEST_RESULTS/ \\
        --out output.xlsx

    python verification_export.py summary \\
        --plan VERIFICATION.md \\
        --results TEST_RESULTS/ \\
        --out summary.md

Dependencies: openpyxl, pyyaml
"""

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

from verification_parser import (
    collect_test_results,
    parse_verification_plan,
)


# ---------------------------------------------------------------------------
# Constants matching the existing Excel layout
# ---------------------------------------------------------------------------

# Column layout (0-indexed logical positions)
# Col 0: spacer
# Col 1: ID, Col 2: Sequence, Col 3: Step, Col 4: Function,
# Col 5: Test, Col 6: Signals, Col 7: Pass/Fail
# Col 8: spacer
# Col 9: Nominal value, Col 10: Unit, Col 11: Tolerance
# Col 12: spacer
# Then per-serial blocks of 7 cols + 1 spacer each

COMMON_HEADERS = [
    "",         # 0: spacer
    "ID",       # 1
    "Sequence", # 2
    "Step",     # 3
    "Function", # 4
    "Test",     # 5
    "Signals",  # 6
    "Pass/Fail",# 7
    "",         # 8: spacer
    "Nominal value",  # 9
    "Unit",     # 10
    "Tolerance",# 11
    "",         # 12: spacer
]

SERIAL_SUBHEADERS = [
    "Measured DMM",
    "Measured ADC Min",
    "Measured ADC Max",
    "Tolerance Min",
    "Tolerance Max",
    "Pass/Fail",
    "Comment",
]

SPACER_COLS_COMMON = [0, 8, 12]  # 0-indexed
SPACER_WIDTH = 1.0
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
HEADER_FONT = Font(bold=True, size=11)
CATEGORY_FONT = Font(bold=True, size=14)

# Conditional format colors
CF_PASS_FILL = PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")
CF_FAIL_FILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
CF_READY_FILL = PatternFill(start_color="FFBDD7EE", end_color="FFBDD7EE", fill_type="solid")
CF_READY_FONT = Font(color="FF0070C0")
CF_PROGRESS_FILL = PatternFill(start_color="FFF4B084", end_color="FFF4B084", fill_type="solid")
CF_PROGRESS_FONT = Font(color="FFFF0000")
CF_NA_FILL = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")


def _serial_col_offset(serial_index: int) -> int:
    """Return the 0-indexed column of the first sub-column for a serial block."""
    # Serial 0 starts at col 13, each block is 7 data + 1 spacer = 8 cols
    return 13 + serial_index * 8


def _total_columns(dut_slots: int) -> int:
    """Total number of columns including spacers."""
    return 13 + dut_slots * 8


def _parse_tolerance(tol_str: str, nominal: float | None) -> tuple[float | None, float | None]:
    """Parse tolerance string and compute min/max bounds.

    Supports: "5%", "0.1", "±0.1", empty.
    Returns (tolerance_min, tolerance_max) or (None, None).
    """
    if not tol_str or nominal is None:
        return None, None

    tol_str = tol_str.strip().replace("±", "")

    if tol_str.endswith("%"):
        try:
            pct = float(tol_str.rstrip("%")) / 100.0
            return nominal * (1 - pct), nominal * (1 + pct)
        except ValueError:
            return None, None
    else:
        try:
            tol = float(tol_str)
            return nominal - tol, nominal + tol
        except ValueError:
            return None, None


def _parse_float(s: str) -> float | None:
    """Parse a string to float, returning None on failure."""
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_excel(plan_path: str, results_dir: str | None, out_path: str) -> None:
    """Export verification plan and results to an Excel workbook."""
    plan = parse_verification_plan(plan_path)
    metadata = plan["metadata"]
    categories = plan["categories"]
    test_steps = plan["test_steps"]
    dut_slots = plan["dut_slots"]

    # Collect results if provided
    results_by_serial: dict[str, list[dict]] = {}
    if results_dir and Path(results_dir).is_dir():
        results_by_serial = collect_test_results(results_dir)

    # Map serial → index (first N serials found, up to dut_slots)
    serial_list = list(results_by_serial.keys())[:dut_slots]

    # Build per-serial lookup: serial → { test_id → measurement_dict }
    serial_lookup: dict[str, dict[str, dict]] = {}
    for serial in serial_list:
        lookup: dict[str, dict] = {}
        for m in results_by_serial[serial]:
            test_id = m.get("ID", "")
            if test_id:
                lookup[test_id] = m
        serial_lookup[serial] = lookup

    total_cols = _total_columns(dut_slots)

    wb = Workbook()
    ws = wb.active
    ws.title = "DUT"

    # ---- Row 1: Top header ----
    row1 = 1
    row2 = 2

    # Common headers span rows 1-2
    for col_idx, header in enumerate(COMMON_HEADERS):
        cell1 = ws.cell(row=row1, column=col_idx + 1, value=header)
        cell1.font = HEADER_FONT
        cell1.border = THIN_BORDER
        cell1.alignment = Alignment(horizontal="center", vertical="center")
        cell2 = ws.cell(row=row2, column=col_idx + 1)
        cell2.border = THIN_BORDER
        # Merge header cells vertically
        if header:
            ws.merge_cells(
                start_row=row1, start_column=col_idx + 1,
                end_row=row2, end_column=col_idx + 1,
            )

    # Serial block headers
    for si in range(dut_slots):
        start_col = _serial_col_offset(si) + 1  # 1-indexed
        serial_name = serial_list[si] if si < len(serial_list) else f"Serial {si + 1}"

        # Merge row 1 across the 7 sub-columns
        ws.merge_cells(
            start_row=row1, start_column=start_col,
            end_row=row1, end_column=start_col + 6,
        )
        cell = ws.cell(row=row1, column=start_col, value=serial_name)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

        # Sub-headers in row 2
        for sub_idx, sub_header in enumerate(SERIAL_SUBHEADERS):
            cell = ws.cell(row=row2, column=start_col + sub_idx, value=sub_header)
            cell.font = Font(bold=True, size=9)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Spacer column after this block
        spacer_col = start_col + 7
        if spacer_col <= total_cols:
            ws.column_dimensions[get_column_letter(spacer_col)].width = SPACER_WIDTH

    # ---- Set spacer column widths ----
    for sc in SPACER_COLS_COMMON:
        ws.column_dimensions[get_column_letter(sc + 1)].width = SPACER_WIDTH

    # ---- Set common column widths ----
    col_widths = {
        1: 8,    # ID
        2: 20,   # Sequence
        3: 25,   # Step
        4: 12,   # Function
        5: 45,   # Test
        6: 28,   # Signals
        7: 12,   # Pass/Fail
        9: 10,   # Nominal
        10: 8,   # Unit
        11: 10,  # Tolerance
    }
    # Offset by 1 because our col indices are 0-based but col_widths keys are 1-based (matching col_idx+1)
    for col_1indexed, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_1indexed)].width = width

    # ---- Data rows ----
    current_row = 3  # First data row

    for cat in categories:
        cat_id = cat["id"]
        cat_name = cat["name"]
        cat_color = cat.get("color", "#FFFFFF").lstrip("#")
        cat_fill = PatternFill(start_color=f"FF{cat_color}", end_color=f"FF{cat_color}", fill_type="solid")
        slots = cat.get("slots", 4)

        steps = test_steps.get(cat_id, [])

        # Pad to fill all slots
        while len(steps) < slots:
            step_num = len(steps)
            steps.append({"ID": f"{cat_id}.{step_num:02d}"})

        cat_start_row = current_row
        cat_end_row = current_row + len(steps) - 1

        for step_idx, step in enumerate(steps):
            row = current_row + step_idx
            test_id = step.get("ID", "")
            nominal = _parse_float(step.get("Nominal", ""))
            tol_str = step.get("Tolerance", "")
            tol_min, tol_max = _parse_tolerance(tol_str, nominal)

            # Common columns
            ws.cell(row=row, column=2, value=test_id).border = THIN_BORDER
            ws.cell(row=row, column=4, value=step.get("Step", "")).border = THIN_BORDER
            ws.cell(row=row, column=5, value=step.get("Function", "")).border = THIN_BORDER
            ws.cell(row=row, column=6, value=step.get("Test", "")).border = THIN_BORDER
            ws.cell(row=row, column=7, value=step.get("Signals", "")).border = THIN_BORDER
            # Overall Pass/Fail — leave empty (computed from serial results)
            pf_cell = ws.cell(row=row, column=8, value="")
            pf_cell.border = THIN_BORDER
            ws.cell(row=row, column=10, value=nominal if nominal is not None else "").border = THIN_BORDER
            if nominal is not None:
                ws.cell(row=row, column=10).number_format = "0.000"
            ws.cell(row=row, column=11, value=step.get("Unit", "")).border = THIN_BORDER
            ws.cell(row=row, column=12, value=tol_str).border = THIN_BORDER

            # Category fill on all common data cells
            for c in range(1, 13):
                ws.cell(row=row, column=c).fill = cat_fill

            # Serial measurement columns
            for si in range(dut_slots):
                start_col = _serial_col_offset(si) + 1
                serial = serial_list[si] if si < len(serial_list) else ""
                measurement = serial_lookup.get(serial, {}).get(test_id, {})

                dmm = _parse_float(measurement.get("Measured DMM", ""))
                adc_min = _parse_float(measurement.get("Measured ADC Min", ""))
                adc_max = _parse_float(measurement.get("Measured ADC Max", ""))
                pf = measurement.get("Pass/Fail", measurement.get("Pass / Fail", ""))
                comment = measurement.get("Comment", "")

                ws.cell(row=row, column=start_col, value=dmm if dmm is not None else "").border = THIN_BORDER
                ws.cell(row=row, column=start_col + 1, value=adc_min if adc_min is not None else "").border = THIN_BORDER
                ws.cell(row=row, column=start_col + 2, value=adc_max if adc_max is not None else "").border = THIN_BORDER
                ws.cell(row=row, column=start_col + 3, value=tol_min if tol_min is not None else "").border = THIN_BORDER
                ws.cell(row=row, column=start_col + 4, value=tol_max if tol_max is not None else "").border = THIN_BORDER
                ws.cell(row=row, column=start_col + 5, value=pf).border = THIN_BORDER
                ws.cell(row=row, column=start_col + 6, value=comment).border = THIN_BORDER

                # Number formats
                for offset in range(5):
                    ws.cell(row=row, column=start_col + offset).number_format = "0.000"

        # Merge Sequence column for category
        ws.merge_cells(
            start_row=cat_start_row, start_column=3,
            end_row=cat_end_row, end_column=3,
        )
        seq_cell = ws.cell(row=cat_start_row, column=3, value=cat_name)
        seq_cell.font = CATEGORY_FONT
        seq_cell.fill = cat_fill
        seq_cell.alignment = Alignment(horizontal="center", vertical="center")
        seq_cell.border = THIN_BORDER

        current_row = cat_end_row + 1

        # Separator row
        for c in range(1, total_cols + 1):
            ws.cell(row=current_row, column=c).border = THIN_BORDER
        ws.row_dimensions[current_row].height = 10
        current_row += 1

    # ---- Conditional formatting for Pass/Fail columns ----
    pf_columns = [8]  # Overall Pass/Fail
    for si in range(dut_slots):
        pf_columns.append(_serial_col_offset(si) + 6 + 1)  # Per-serial Pass/Fail (1-indexed)

    data_end_row = current_row - 1
    for col_1indexed in pf_columns:
        col_letter = get_column_letter(col_1indexed)
        cell_range = f"{col_letter}3:{col_letter}{data_end_row}"

        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[f'SEARCH("Pass",{col_letter}3)'],
                fill=CF_PASS_FILL,
            ),
        )
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[f'SEARCH("Fail",{col_letter}3)'],
                fill=CF_FAIL_FILL,
            ),
        )
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[f'EXACT({col_letter}3,"N/A")'],
                fill=CF_NA_FILL,
            ),
        )

    # ---- Freeze panes ----
    ws.freeze_panes = "D3"

    wb.save(out_path)
    print(f"Excel exported to: {out_path}")


# ---------------------------------------------------------------------------
# Markdown summary export
# ---------------------------------------------------------------------------

def export_summary(plan_path: str, results_dir: str | None, out_path: str) -> None:
    """Export a Markdown summary report."""
    plan = parse_verification_plan(plan_path)
    metadata = plan["metadata"]
    categories = plan["categories"]
    test_steps = plan["test_steps"]

    results_by_serial: dict[str, list[dict]] = {}
    if results_dir and Path(results_dir).is_dir():
        results_by_serial = collect_test_results(results_dir)

    # Flatten all measurements into a lookup by test ID
    all_measurements: dict[str, list[dict]] = {}
    for serial, measurements in results_by_serial.items():
        for m in measurements:
            tid = m.get("ID", "")
            if tid:
                all_measurements.setdefault(tid, []).append(m)

    lines = []
    lines.append(f"# Verification Summary — {metadata.get('product', '')} {metadata.get('revision', '')}")
    lines.append("")
    lines.append(f"**Phase:** {metadata.get('phase', '')}")
    lines.append(f"**Engineer:** {metadata.get('engineer', '')}")
    lines.append(f"**Date:** {metadata.get('date_created', '')}")
    lines.append(f"**DUTs tested:** {len(results_by_serial)}")
    lines.append("")
    lines.append("---")
    lines.append("")

    # Per-category summary
    total_defined = 0
    total_pass = 0
    total_fail = 0
    total_open = 0

    lines.append("## Results by Category")
    lines.append("")
    lines.append("| Category | Defined | Pass | Fail | Open |")
    lines.append("|----------|---------|------|------|------|")

    for cat in categories:
        cat_id = cat["id"]
        cat_name = cat["name"]
        steps = test_steps.get(cat_id, [])
        # Only count defined steps (non-empty Step or Test field)
        defined = [s for s in steps if s.get("Step") or s.get("Test")]

        cat_pass = 0
        cat_fail = 0
        cat_open = 0
        for s in defined:
            tid = s.get("ID", "")
            meas = all_measurements.get(tid, [])
            if not meas:
                cat_open += 1
            elif any(m.get("Pass/Fail", "").lower().startswith("fail") for m in meas):
                cat_fail += 1
            elif all(m.get("Pass/Fail", "").lower().startswith("pass") for m in meas):
                cat_pass += 1
            else:
                cat_open += 1

        lines.append(f"| {cat_name} | {len(defined)} | {cat_pass} | {cat_fail} | {cat_open} |")
        total_defined += len(defined)
        total_pass += cat_pass
        total_fail += cat_fail
        total_open += cat_open

    lines.append(f"| **Total** | **{total_defined}** | **{total_pass}** | **{total_fail}** | **{total_open}** |")
    lines.append("")

    # Failures detail
    lines.append("---")
    lines.append("")
    lines.append("## Failures")
    lines.append("")

    has_failures = False
    for cat in categories:
        cat_id = cat["id"]
        steps = test_steps.get(cat_id, [])
        for s in steps:
            tid = s.get("ID", "")
            meas = all_measurements.get(tid, [])
            for m in meas:
                if m.get("Pass/Fail", "").lower().startswith("fail"):
                    has_failures = True
                    lines.append(f"- **{tid}** {s.get('Step', '')}: {m.get('Comment', 'No comment')}")

    if not has_failures:
        lines.append("_No failures recorded._")

    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## DUT Summary")
    lines.append("")
    if results_by_serial:
        lines.append("| Serial | Tests Recorded |")
        lines.append("|--------|---------------|")
        for serial, measurements in results_by_serial.items():
            lines.append(f"| {serial} | {len(measurements)} |")
    else:
        lines.append("_No test results recorded yet._")

    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("> Convert to PDF: `pandoc summary.md -o summary.pdf`")
    lines.append("> Or open in a browser / VS Code and print to PDF.")

    Path(out_path).write_text("\n".join(lines), encoding="utf-8")
    print(f"Summary exported to: {out_path}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export verification plans to Excel or Markdown summary.",
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    # excel subcommand
    p_excel = subparsers.add_parser("excel", help="Export to Excel (.xlsx)")
    p_excel.add_argument("--plan", required=True, help="Path to VERIFICATION.md")
    p_excel.add_argument("--results", default=None, help="Path to TEST_RESULTS/ directory")
    p_excel.add_argument("--out", required=True, help="Output .xlsx path")

    # summary subcommand
    p_summary = subparsers.add_parser("summary", help="Export Markdown summary")
    p_summary.add_argument("--plan", required=True, help="Path to VERIFICATION.md")
    p_summary.add_argument("--results", default=None, help="Path to TEST_RESULTS/ directory")
    p_summary.add_argument("--out", required=True, help="Output .md path")

    args = parser.parse_args()

    if args.command == "excel":
        export_excel(args.plan, args.results, args.out)
    elif args.command == "summary":
        export_summary(args.plan, args.results, args.out)


if __name__ == "__main__":
    main()
