import openpyxl

FILE = r"20_Projects\ESH10000543\R2\DOCS\DesignLog_FixttureLink.xlsx"
TAB = "Rev 1"

wb = openpyxl.load_workbook(FILE, data_only=True)
print(f"Sheets: {wb.sheetnames}\n")

if TAB not in wb.sheetnames:
    print(f"Tab '{TAB}' not found.")
else:
    ws = wb[TAB]
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            print("\t".join("" if c is None else str(c) for c in row))
